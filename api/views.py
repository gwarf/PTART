from django.contrib.auth import authenticate
from django.contrib.auth.password_validation import validate_password
from django.core.exceptions import ValidationError

from django.http import HttpResponse

from django.views.decorators.csrf import csrf_exempt

from rest_framework import status
from rest_framework.decorators import action
from rest_framework.decorators import api_view
from rest_framework.renderers import BaseRenderer
from rest_framework.response import Response
from rest_framework.authtoken.models import Token

from openpyxl import Workbook, styles
from openpyxl.writer.excel import save_virtual_workbook

from svglib.svglib import svg2rlg
from reportlab.graphics import renderPM

import jinja2
import json
import os
import pypandoc
import random
import re
import zipfile

from ptart.models import Flag, Hit, Assessment, Project, Template, Comment, HitReference, Host, Service, Screenshot, Attachment, Cvss, Case, Module, Methodology, Label, AttackScenario

from api.decorators import ptart_authentication

from .serializers import FlagSerializer, HitSerializer, AssessmentSerializer, ProjectSerializer, TemplateSerializer, HostSerializer, ServiceSerializer, ScreenshotSerializer, AttachmentSerializer, CommentSerializer, HitReferenceSerializer, CvssSerializer, CaseSerializer, ModuleSerializer, MethodologySerializer, LabelSerializer, AttackScenarioSerializer

@csrf_exempt
@ptart_authentication
@api_view(['GET', 'PATCH', 'PUT', 'DELETE'])
def flag(request, pk):
    return item(request, pk, Flag, FlagSerializer)

@csrf_exempt
@ptart_authentication
@api_view(['GET', 'POST'])
def flags(request):
    return items(request, Flag, FlagSerializer)

@csrf_exempt
@ptart_authentication
@api_view(['GET', 'PATCH', 'PUT', 'DELETE'])
def hit(request, pk):
    return item(request, pk, Hit, HitSerializer)

@csrf_exempt
@ptart_authentication
@api_view(['GET', 'POST'])
def hits(request):
    return items(request, Hit, HitSerializer)

@csrf_exempt
@ptart_authentication
@api_view(['GET', 'PATCH', 'PUT', 'DELETE'])
def attackscenario(request, pk):
    return item(request, pk, AttackScenario, AttackScenarioSerializer)

@csrf_exempt
@ptart_authentication
@api_view(['GET', 'POST'])
def attackscenarios(request):
    return items(request, AttackScenario, AttackScenarioSerializer)

@csrf_exempt
@ptart_authentication
@api_view(['GET', 'PATCH', 'PUT', 'DELETE'])
def label(request, pk):
    return item(request, pk, Label, LabelSerializer)

@csrf_exempt
@ptart_authentication
@api_view(['GET', 'POST'])
def labels(request):
    return items(request, Label, LabelSerializer)

@csrf_exempt
@ptart_authentication
@api_view(['GET', 'DELETE'])
def comment(request, pk):
    return item(request, pk, Comment, CommentSerializer)

@csrf_exempt
@ptart_authentication
@api_view(['GET', 'DELETE'])
def hit_reference(request, pk):
    return item(request, pk, HitReference, HitReferenceSerializer)

@csrf_exempt
@ptart_authentication
@api_view(['DELETE'])
def screenshot(request, pk):
    return item(request, pk, Screenshot, ScreenshotSerializer)

@csrf_exempt
@ptart_authentication
@api_view(['GET', 'POST'])
def screenshots(request):
    return items(request, Screenshot, ScreenshotSerializer)

@csrf_exempt
@ptart_authentication
@api_view(['DELETE'])
def attachment(request, pk):
    return item(request, pk, Attachment, AttachmentSerializer)

@csrf_exempt
@ptart_authentication
@api_view(['GET', 'POST'])
def attachments(request):
    return items(request, Attachment, AttachmentSerializer)

@csrf_exempt
@ptart_authentication
@api_view(['GET', 'PATCH', 'PUT', 'DELETE'])
def assessment(request, pk):
    return item(request, pk, Assessment, AssessmentSerializer)

@csrf_exempt
@ptart_authentication
@api_view(['GET', 'POST'])
def assessments(request):
    return items(request, Assessment, AssessmentSerializer)

@csrf_exempt
@ptart_authentication
@api_view(['GET', 'PATCH', 'PUT', 'DELETE'])
def host(request, pk):
    return item(request, pk, Host, HostSerializer)

@csrf_exempt
@ptart_authentication
@api_view(['GET', 'POST'])
def hosts(request):
    return items(request, Host, HostSerializer)

@csrf_exempt
@ptart_authentication
@api_view(['GET', 'PATCH', 'PUT', 'DELETE'])
def service(request, pk):
    return item(request, pk, Service, ServiceSerializer)

@csrf_exempt
@ptart_authentication
@api_view(['GET', 'POST'])
def services(request):
    return items(request, Service, ServiceSerializer)

@csrf_exempt
@ptart_authentication
@api_view(['GET', 'PATCH', 'PUT', 'DELETE'])
def project(request, pk):
    return item(request, pk, Project, ProjectSerializer)

@csrf_exempt
@ptart_authentication
@api_view(['GET', 'POST'])
def projects(request):
    return items(request, Project, ProjectSerializer)

@csrf_exempt
@ptart_authentication
@api_view(['GET', 'PATCH', 'PUT', 'DELETE'])
def template(request, pk):
    return item(request, pk, Template, TemplateSerializer)

@csrf_exempt  
@ptart_authentication
@api_view(['GET', 'POST'])
def templates(request):
    return items(request, Template, TemplateSerializer)

@csrf_exempt
@ptart_authentication
@api_view(['GET', 'PATCH', 'PUT', 'DELETE'])
def case(request, pk):
    return item(request, pk, Case, CaseSerializer)

@csrf_exempt
@ptart_authentication
@api_view(['GET', 'POST'])
def cases(request):
    return items(request, Case, CaseSerializer)

@csrf_exempt
@ptart_authentication
@api_view(['GET', 'PATCH', 'PUT', 'DELETE'])
def module(request, pk):
    return item(request, pk, Module, ModuleSerializer)

@csrf_exempt
@ptart_authentication
@api_view(['GET', 'POST'])
def modules(request):
    return items(request, Module, ModuleSerializer)

@csrf_exempt
@ptart_authentication
@api_view(['GET', 'PATCH', 'PUT', 'DELETE'])
def methodology(request, pk):
    return item(request, pk, Methodology, MethodologySerializer)

@csrf_exempt
@ptart_authentication
@api_view(['GET', 'POST'])
def methodologies(request):
    return items(request, Methodology, MethodologySerializer)

@csrf_exempt
@ptart_authentication
@api_view(['POST'])
def load_module(request, pk, assessmentId):
    response = None
    try:
        assessment = Assessment.objects.get(pk=assessmentId)
        if assessment.is_user_can_edit(request.user) :
            module = Module.objects.get(pk=pk)
            cases = Case.objects.filter(module=module)
            flags = []
            for case in cases:
                note = "Module: " + module.name + "\n\n" + case.description
                flag = Flag(title=case.name, note=note, assessment=assessment, assignee = request.user)
                flag.save()
                flags.append(flag)
            response = Response(FlagSerializer(flags, many=True).data, status=status.HTTP_201_CREATED)
        else :
            response = Response(status=status.HTTP_403_FORBIDDEN)
    except Module.DoesNotExist:
        response = Response(status=status.HTTP_404_NOT_FOUND)
    except Assessment.DoesNotExist:
        response = Response(status=status.HTTP_404_NOT_FOUND)
    return response

@csrf_exempt
@ptart_authentication
@action(methods=['GET'], detail=True)
def screenshot_raw(request, pk) :
    response = None
    try:
        item = Screenshot.objects.get(pk=pk)
        if item.is_user_can_view(request.user) :
            response = Response(item.get_raw_data())        
        else :
            response = Response(status=status.HTTP_403_FORBIDDEN)
    except Screenshot.DoesNotExist:
        response = Response(status=status.HTTP_404_NOT_FOUND)
        
    response.accepted_renderer = ImageRenderer()
    response.accepted_media_type = 'image/png'
    response.renderer_context = {}
    return response

@csrf_exempt
@ptart_authentication
@action(methods=['GET'], detail=True)
def attachment_raw(request, pk) :
    response = None
    try:
        item = Attachment.objects.get(pk=pk)
        if item.is_user_can_view(request.user) :
            response = Response(item.get_raw_data())
            response.content_type = "application/octet-stream"
            response['Content-Disposition'] = 'attachment; filename=' + item.attachment_name
        else :
            response = Response(status=status.HTTP_403_FORBIDDEN)
    except Attachment.DoesNotExist:
        response = Response(status=status.HTTP_404_NOT_FOUND)
        
    response.accepted_renderer = BinaryRenderer()
    response.accepted_media_type = 'application/octet-stream'
    response.renderer_context = {}
    return response

@csrf_exempt
@ptart_authentication
@api_view(['POST'])
def cvss(request):
    serializer = CvssSerializer(data=request.data)
    if serializer.is_valid():
        cvss = Cvss(**serializer.validated_data)
        cvss.compute_cvss_value()
        response = Response(CvssSerializer(cvss).data, status=status.HTTP_201_CREATED)
    else :
        response = Response(serializer.errors, status=status.HTTP_400_BAD_REQUEST)
    return response

@csrf_exempt
@ptart_authentication
@api_view(['POST','DELETE'])
def cvss_hit(request, pk):
    response = None
    try:
        hit = Hit.objects.get(pk=pk)
        if hit.is_user_can_edit(request.user):
            
            if request.method == 'DELETE':
                #Delete the cvss attached to the hit.
                if hit.cvss is not None :
                    Cvss.objects.get(pk=hit.cvss.id).delete()
                response = Response(status=status.HTTP_200_OK)

            else :
                serializer = CvssSerializer(data=request.data)
                if serializer.is_valid():
                    #Create the new CVSS.
                    serializer.save()
                    cvss = Cvss.objects.get(pk=serializer.data["id"])
                    cvss.compute_cvss_value()
                    cvss.save(update_fields=['decimal_value'])

                    #This condition prevent memory leak in DB.
                    if hit.cvss is not None : 
                        Cvss.objects.get(pk=hit.cvss.id).delete()

                    hit.cvss = cvss
                    hit.save(update_fields=['cvss'])
                    response = Response(status=status.HTTP_201_CREATED)
                else :
                    response = Response(serializer.errors, status=status.HTTP_400_BAD_REQUEST) 
        else :
            response = Response(status=status.HTTP_403_FORBIDDEN)
    except Hit.DoesNotExist:
        response = Response(status=status.HTTP_404_NOT_FOUND)
    return response

@csrf_exempt
@ptart_authentication
@api_view(['POST','GET'])
def comments(request, pk):
    response = None
    try:
        hit = Hit.objects.get(pk=pk)
        if request.method == 'GET':
            if hit.is_user_can_view(request.user):
                response = Response(CommentSerializer(hit.comment_set.all(), many=True).data)
            else :
                response = Response(status=status.HTTP_403_FORBIDDEN)
        else :
            text = request.data["text"]
            if text is not None and text.strip() :
                try:
                    comment = Comment(text=text, hit=hit, author = request.user)
                    if comment.is_user_can_create(request.user):
                        comment.save()
                        response = Response(CommentSerializer(comment).data, status=status.HTTP_200_OK)
                    else :
                        response = Response(status=status.HTTP_403_FORBIDDEN)
                except Hit.DoesNotExist:
                    response = Response(status=status.HTTP_404_NOT_FOUND)
            else :
                response = Response(status=status.HTTP_400_BAD_REQUEST)
    except Hit.DoesNotExist:
        response = Response(status=status.HTTP_404_NOT_FOUND)
    return response


@csrf_exempt
@ptart_authentication
@api_view(['POST','GET'])
def hit_references(request, pk):
    response = None
    try:
        hit = Hit.objects.get(pk=pk)
        if request.method == 'GET':
            if hit.is_user_can_view(request.user):
                response = Response(HitReferenceSerializer(hit.hitreference_set.all(), many=True).data)
            else :
                response = Response(status=status.HTTP_403_FORBIDDEN)
        else :
            name = request.data["name"]
            url = request.data["url"]
            if name is not None and name.strip() and url is not None and url.strip() and url.startswith("http") :
                try:
                    hitreference = HitReference(name=name, url=url, hit=hit)
                    if hitreference.is_user_can_create(request.user):
                        hitreference.save()
                        response = Response(HitReferenceSerializer(hitreference).data, status=status.HTTP_200_OK)
                    else :
                        response = Response(status=status.HTTP_403_FORBIDDEN)
                except Hit.DoesNotExist:
                    response = Response(status=status.HTTP_404_NOT_FOUND)
            else :
                response = Response(status=status.HTTP_400_BAD_REQUEST)
    except Hit.DoesNotExist:
        response = Response(status=status.HTTP_404_NOT_FOUND)
    return response

@csrf_exempt
@ptart_authentication
@api_view(['POST'])
def markFlagAsDone(request, pk) :
    response = None
    try:
        flag = Flag.objects.get(pk=pk)
        if flag.is_user_can_edit(request.user):
            flag.done = True
            flag.save()
            response = Response(FlagSerializer(flag).data, status=status.HTTP_200_OK)
        else :
            response = Response(status=status.HTTP_403_FORBIDDEN)
    except Flag.DoesNotExist:
        response = Response(status=status.HTTP_404_NOT_FOUND)
    return response

@csrf_exempt
@ptart_authentication
@api_view(['GET'])
def project_burp_configuration(request, pk):
    response = None
    try:
        project = Project.objects.get(pk=pk)
        if project.is_user_can_view(request.user):
            configuration = {"target": { "scope": { "advanced_mode": "true"}}}
            targets = []
            for target in project.scope.splitlines() :
                burp_target = {}
                burp_target["enabled"] = True

                #Manage protocol
                protocol = "any"
                if target.startswith("https://") :
                    protocol = "https"
                    target = target.replace("https://", "")
                elif target.startswith("http://") :
                    protocol = "http"
                    target = target.replace("http://", "")
                burp_target["protocol"] = protocol

                #Manage port
                if ":" in target :
                    port = target.split(":")[1]
                    if "/" in port :
                        port = port.split("/")[0]
                    burp_target["port"] = "^"+port + "$"
                    target = target.replace(":" + port, "")

                #Manage file
                if "/" in target :
                    file = target.partition("/")[2]
                    burp_target["file"] = "^" + file + "$"
                    target = target.replace("/" + file, "")

                #Manage Target
                target = target.replace(".", "\.")
                target = target.replace("*", ".*")
                burp_target["host"] = "^" + target + "$"

                targets.append(burp_target)
            configuration["target"]["scope"]["include"] = targets
            response = Response(configuration)
            response.accepted_renderer = JsonRenderer()
            response.accepted_media_type = 'application/json'
            response.renderer_context = {}
            response['Content-Disposition'] = 'attachment; filename=' + project.name + "_burp_configuration.json"
        else :
            response = Response(status=status.HTTP_403_FORBIDDEN)
    except Flag.DoesNotExist:
        response = Response(status=status.HTTP_404_NOT_FOUND)
    return response

@csrf_exempt
@ptart_authentication
@action(methods=['GET'], detail=True)
def project_xlsx(request, pk):
    response = None
    try:
        project = Project.objects.get(pk=pk)
        if project.is_user_can_view(request.user):
                        
            wb = Workbook()
            ws = wb.active

            #Define column size
            wb.active.column_dimensions['A'].width = 28
            wb.active.column_dimensions['B'].width = 10
            wb.active.column_dimensions['C'].width = 10
            wb.active.column_dimensions['D'].width = 17
            wb.active.column_dimensions['E'].width = 50
            wb.active.column_dimensions['F'].width = 30
            wb.active.column_dimensions['G'].width = 12
            wb.active.column_dimensions['H'].width = 50

            #Add project data.
            ws['A1'] = "Project Name:"
            ws['A2'] = "Client:"
            ws['A3'] = "Date:"
            ws['A4'] = "Auditors:"
            ws['B1'] = project.name
            ws['B2'] = project.client
            if project.start_date is not None and project.end_date is not None :
                ws['B3'] = "From " + str(project.start_date) + " To " + str(project.end_date)
            else :
                ws['B3'] = project.added

            
            #Construct the auditor string
            pentester_str = ""
            previous = ""
            for pentester in project.pentesters.all():
                pentester_str = "{}{}{} - {} {}".format(pentester_str, previous, pentester.username, pentester.first_name, pentester.last_name)
                previous = ", "
            ws['B3'] = pentester_str

            #Beautify project data 
            ws.merge_cells('B1:H1')
            ws.merge_cells('B2:H2')
            ws.merge_cells('B3:H3')
            ws.merge_cells('B4:H4')

            
            projectHeaderStyle = styles.NamedStyle(name = 'project_header_style')
            projectHeaderStyle.font = styles.Font(name = 'Calibri', size = 14, bold = True, color = '000000')
            projectHeaderStyle.fill = styles.PatternFill(patternType = 'solid', fgColor = '00B0F0')
            projectHeaderStyle.alignment = styles.Alignment(horizontal= 'left')
            ws['A1'].style = projectHeaderStyle
            ws['A2'].style = projectHeaderStyle
            ws['A3'].style = projectHeaderStyle
            ws['A4'].style = projectHeaderStyle

            projectValueStyle = styles.NamedStyle(name = 'project_value_style')
            projectValueStyle.font = styles.Font(name = 'Calibri', size = 14, italic = True, color = '000000')
            projectValueStyle.fill = styles.PatternFill(patternType = 'solid', fgColor = '00B0F0')
            projectValueStyle.alignment = styles.Alignment(horizontal= 'left')

            ws['B1'].style = projectValueStyle
            ws['B2'].style = projectValueStyle
            ws['B3'].style = projectValueStyle
            ws['B4'].style = projectValueStyle
            ws['B3'].number_format = 'YYYY MMM DD'


            #Add column header.
            ws['A6'] = "Assessment"
            ws['B6'] = "Sev"
            ws['C6'] = "CVSS"
            ws['D6'] = "ID"
            ws['E6'] = "Title"
            ws['F6'] = "Asset"
            ws['G6'] = "Fix Compl."
            ws['H6'] = "Labels"

            columnHeaderStyle = styles.NamedStyle(name = 'column_header_style')
            columnHeaderStyle.font = styles.Font(name = 'Calibri', size = 12, bold = True, color = '000000')
            columnHeaderStyle.fill = styles.PatternFill(patternType = 'solid', fgColor = '92D050')
            columnHeaderStyle.alignment = styles.Alignment(horizontal= 'center')

            ws['A6'].style = columnHeaderStyle
            ws['B6'].style = columnHeaderStyle
            ws['C6'].style = columnHeaderStyle
            ws['D6'].style = columnHeaderStyle
            ws['E6'].style = columnHeaderStyle
            ws['F6'].style = columnHeaderStyle
            ws['G6'].style = columnHeaderStyle
            ws['H6'].style = columnHeaderStyle

            #Fill the report
            criticalStyle = styles.NamedStyle(name = 'critical_style')
            criticalStyle.font = styles.Font(name = 'OCR A Extended', color = 'FFFFFF')
            criticalStyle.fill = styles.PatternFill(patternType = 'solid', fgColor = '343a40')
            criticalStyle.alignment = styles.Alignment(horizontal= 'center')

            highStyle = styles.NamedStyle(name = 'high_style')
            highStyle.font = styles.Font(name = 'OCR A Extended', color = 'FFFFFF')
            highStyle.fill = styles.PatternFill(patternType = 'solid', fgColor = 'dc3545')
            highStyle.alignment = styles.Alignment(horizontal= 'center')

            mediumStyle = styles.NamedStyle(name = 'medium_style')
            mediumStyle.font = styles.Font(name = 'OCR A Extended', color = '212529')
            mediumStyle.fill = styles.PatternFill(patternType = 'solid', fgColor = 'ffc107')
            mediumStyle.alignment = styles.Alignment(horizontal= 'center')

            lowStyle = styles.NamedStyle(name = 'low_style')
            lowStyle.font = styles.Font(name = 'OCR A Extended', color = 'FFFFFF')
            lowStyle.fill = styles.PatternFill(patternType = 'solid', fgColor = '28a745')
            lowStyle.alignment = styles.Alignment(horizontal= 'center')

            infoStyle = styles.NamedStyle(name = 'info_style')
            infoStyle.font = styles.Font(name = 'OCR A Extended', color = 'FFFFFF')
            infoStyle.fill = styles.PatternFill(patternType = 'solid', fgColor = '6c757d')
            infoStyle.alignment = styles.Alignment(horizontal= 'center')

            line = 7
            for assessment in project.assessment_set.all():        
                for hit in assessment.displayable_hits():
                    ws.cell(row=line, column=1).value = assessment.name
                    ws.cell(row=line, column=2).value = "P{}".format(hit.severity)
                    ws.cell(row=line, column=3).value = hit.get_cvss_value()
                    ws.cell(row=line, column=4).value = hit.get_unique_id()
                    ws.cell(row=line, column=5).value = hit.title
                    ws.cell(row=line, column=6).value = hit.asset
                    ws.cell(row=line, column=7).value = hit.get_fix_complexity_str()

                    label_str = ""
                    previous = ""
                    for label in hit.labels.all():
                        label_str = "{}{}{}".format(label_str, previous, label.title)
                        previous = ", "
                    ws.cell(row=line, column=8).value = label_str

                    #Apply style from value.
                    if hit.severity == 1:
                        ws.cell(row=line, column=2).style = criticalStyle
                    elif hit.severity == 2:
                        ws.cell(row=line, column=2).style = highStyle
                    elif hit.severity == 3:
                        ws.cell(row=line, column=2).style = mediumStyle
                    elif hit.severity == 4:
                        ws.cell(row=line, column=2).style = lowStyle
                    elif hit.severity == 5:
                        ws.cell(row=line, column=2).style = infoStyle

                    if hit.fix_complexity == 1:
                        ws.cell(row=line, column=7).style = highStyle
                    elif hit.fix_complexity == 2:
                        ws.cell(row=line, column=7).style = mediumStyle
                    elif hit.fix_complexity == 3:
                        ws.cell(row=line, column=7).style = lowStyle
                    else :
                        ws.cell(row=line, column=7).style = infoStyle

                    try:
                        if float(hit.get_cvss_value()) < 4.0:
                            ws.cell(row=line, column=3).style = infoStyle
                        elif float(hit.get_cvss_value()) < 4.0:
                            ws.cell(row=line, column=3).style = lowStyle
                        elif float(hit.get_cvss_value()) < 7.0:
                            ws.cell(row=line, column=3).style = mediumStyle
                        elif float(hit.get_cvss_value()) < 9.0:
                            ws.cell(row=line, column=3).style = highStyle
                        else :
                            ws.cell(row=line, column=3).style = criticalStyle
                    except ValueError:
                        ws.cell(row=line, column=3).style = infoStyle

                    line = line + 1
            
            ws.auto_filter.ref = "A6:H{}".format(line)

            #Prepare HTTP response.
            response = Response(save_virtual_workbook(wb))
            response.content_type = 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
            response['Content-Disposition'] = 'attachment; filename=' + project.name + ".xlsx"
            response.accepted_renderer = BinaryRenderer()
            response.accepted_media_type = 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
            response.renderer_context = {}

            wb.close()
        else :
            response = Response(status=status.HTTP_403_FORBIDDEN)
    except Flag.DoesNotExist:
        response = Response(status=status.HTTP_404_NOT_FOUND)
    return response

@csrf_exempt
@ptart_authentication
@action(methods=['GET'], detail=True)
def project_latex(request, pk):
    response = None
    try:
        project = Project.objects.get(pk=pk)
        if project.is_user_can_view(request.user):
            response = HttpResponse()
            zf = zipfile.ZipFile(response, 'w')
            
            #Retrieve Screenshots.
            for assessment in project.assessment_set.all():        
                for hit in assessment.displayable_hits():
                    for screenshot in hit.screenshot_set.all():
                        zf.writestr("screenshots/{}.png".format(screenshot.id), screenshot.get_raw_data())
            
            #Retrieve AttackScenarios.
            for attackscenario in project.attackscenario_set.all():
                #SVG must be converted to PNG to facilitate integration in LaTeX report.
                tempSvgFilename = "{}_{}_{}.svg".format(project.id, attackscenario.id,random.randint(0,10000))
                tempPngFilename = "{}_{}_{}.png".format(project.id, attackscenario.id,random.randint(0,10000))
                f = open(tempSvgFilename, "a")
                f.write(attackscenario.svg)
                f.close()
                drawing = svg2rlg(tempSvgFilename)
                renderPM.drawToFile(drawing, tempPngFilename, fmt="PNG")    
                zf.write(tempPngFilename, "attackscenarios/{}.png".format(attackscenario.id))
                os.remove(tempSvgFilename)
                os.remove(tempPngFilename)

            #Add resources for LaTeX
            zf.write("reports/resources/companylogo.png", "resources/companylogo.png")
            zf.write("reports/resources/logo.png", "resources/logo.png")
            
            #Generate Latex report.
            #Custom environment is used to avoid syntax conflict between Jinja & LaTex
            with open('reports/report_latex.tex') as file_:
                env = jinja2.Environment(
                    block_start_string = '\BLOCK{',
                    block_end_string = '}',
                    variable_start_string = '\VAR{',
                    variable_end_string = '}',
                    comment_start_string = '\#{',
                    comment_end_string = '}',
                    line_statement_prefix = '%%',
                    line_comment_prefix = '%#',
                    trim_blocks = True,
                    autoescape = False,
                    loader = jinja2.FileSystemLoader(os.path.abspath('.'))
                )
                
                #Filters that has been used in the LaTex Report.
                escape_tex_table = {
                        '&': r'\&',
                        '%': r'\%',
                        '$': r'\$',
                        '#': r'\#',
                        '_': r'\_',
                        '~': r'\textasciitilde{}',
                        '^': r'\^{}',
                        '\\': r'\textbackslash{}',
                        '<': r'\textless{}',
                        '>': r'\textgreater{}',
                    }
                tex_regex = re.compile('|'.join(re.escape(str(key)) for key in sorted(escape_tex_table.keys(), key = lambda item: - len(item))))
                
                def tex_escape(text):
                    return tex_regex.sub(lambda match: escape_tex_table[match.group()], text)

                def markdown_to_latex(md) :
                    return pypandoc.convert_text(tex_escape(md), 'latex', format='md', extra_args=['--wrap=preserve', '--highlight-style=tango'])
                #End of filters.

                env.filters["mdtolatex"] = markdown_to_latex
                env.filters["escape"] = tex_escape
                template = env.from_string(file_.read())
                zf.writestr("report.tex", template.render(project=project, labels=Label.get_viewable(request.user)))     

            #Prepare HTTP response.
            response.content_type = 'application/zip'
            response['Content-Disposition'] = 'attachment; filename=' + project.name + ".zip"
            response.accepted_renderer = BinaryRenderer()
            response.accepted_media_type = 'application/zip'
            response.renderer_context = {}
        else :
            response = Response(status=status.HTTP_403_FORBIDDEN)
    except Flag.DoesNotExist:
        response = Response(status=status.HTTP_404_NOT_FOUND)
    return response

@csrf_exempt
@ptart_authentication
@api_view(['POST','DELETE'])
def manage_token(request):
    """
        Grant or Revoke authentication token.
    """
    response = Response(status=status.HTTP_404_NOT_FOUND)
    if request.method == 'POST':
        response = Response(status=status.HTTP_404_NOT_FOUND)
        if not Token.objects.filter(user=request.user).exists():
            token = Token.objects.create(user=request.user)
            response = Response({"token" : token.key}, status=status.HTTP_201_CREATED)
    else :
        response = Response(status=status.HTTP_404_NOT_FOUND)
        if Token.objects.filter(user=request.user).exists():
            Token.objects.filter(user=request.user).delete()
            response = Response({"token" : ""}, status=status.HTTP_204_NO_CONTENT)

    return response

@csrf_exempt
@ptart_authentication
@api_view(['POST'])
def change_password(request):
    """
        Change current user password.
    """
    response = Response({"The current password is not valid"}, status=status.HTTP_400_BAD_REQUEST)
    user = authenticate(request, username=request.user, password=request.data["oldPassword"])
    if user is not None:
        password1 = str(request.data["newPassword1"])
        password2 = str(request.data["newPassword2"])
        if password1 == password2 :
            try :
                validate_password(password1, user)
                user.set_password(password1)
                user.save()
                response = Response({}, status=status.HTTP_204_NO_CONTENT)
            except ValidationError as err:
                response = Response(err, status=status.HTTP_400_BAD_REQUEST)
        else :
            response = Response({"New passwords are not matching"}, status=status.HTTP_400_BAD_REQUEST)
    return response
 
#
# CRUD operations for a specific item.
#
def item(request, pk, class_name, serializer_name) :
    response = None
    try:
        item = class_name.objects.get(pk=pk)

        if request.method == 'GET':
            if item.is_user_can_view(request.user) :
                response = Response(serializer_name(item).data)
            else :
                response = Response(status=status.HTTP_403_FORBIDDEN)
        elif request.method == 'PUT' or request.method == 'PATCH':
            serializer = serializer_name(item, data=request.data)
            if serializer.is_valid():
                if item.is_user_can_edit(request.user) :
                    serializer.save()
                    response = Response(serializer.data)
                else :
                    response = Response(status=status.HTTP_403_FORBIDDEN)
            else :
                response = Response(serializer.errors, status=status.HTTP_400_BAD_REQUEST)
        
        elif request.method == 'DELETE':
            if item.is_user_can_edit(request.user) :
                response = Response(serializer_name(item).data, status=status.HTTP_200_OK)
                item.delete()
            else :
                response = Response(status=status.HTTP_403_FORBIDDEN)

    except class_name.DoesNotExist:
        response = Response(status=status.HTTP_404_NOT_FOUND)

    return response

#
# CRUD operations for a specific list of items.
#
def items(request, class_name, serializer_name) :
    response = None
    if request.method == 'GET':
        response = Response(serializer_name(class_name.get_viewable(request.user), many=True).data)
    elif request.method == 'POST':
        serializer = serializer_name(data=request.data)
        if serializer.is_valid():
            if class_name(**serializer.validated_data).is_user_can_create(request.user) :
                serializer.save()
                response = Response(serializer.data, status=status.HTTP_201_CREATED)
            else :
                response = Response(status=status.HTTP_403_FORBIDDEN)
        else :
            response = Response(serializer.errors, status=status.HTTP_400_BAD_REQUEST)
    
    return response

#
# Image renderer for Screenshots.
#
class ImageRenderer(BaseRenderer):
    def render(self, data, media_type='image/png', renderer_context=None):
        return data

#
# JSON renderer for Burp config file.
#
class JsonRenderer(BaseRenderer):
    def render(self, data, media_type='application/json', renderer_context=None):
        return data

#
# Binary renderer for Attachments.
#
class BinaryRenderer(BaseRenderer):
    def render(self, data, media_type='application/octet-stream', renderer_context=None):
        return data
